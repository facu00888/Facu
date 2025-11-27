# -*- coding: utf-8 -*-
"""
invoice_parser.py
- Parser de facturas (principalmente PDF con capa de texto) para Uruguay (DGI e-Factura típicas).

Diseñado para:
- NO depender del nombre del archivo (eso no escala).
- Priorizar extracción desde PDF-text (pdfplumber) y evitar OCR salvo que lo pidas.
- Sacar campos comunes: fecha, serie, folio, rut emisor/receptor, totales, etc.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


# -----------------------------
# Utilidades generales
# -----------------------------

UY_DATE_RE = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{4})\b")
RUT_RE = re.compile(r"\b(\d{12})\b")

BUSINESS_HINT_RE = re.compile(
    r"\b(SRL|S\.R\.L\.|SA|S\.A\.|SAS|S\.A\.S\.|LTDA|Ltda|SCS|S\.C\.S\.|S\.A\.U\.|S\.A\.P\.A\.)\b",
    re.IGNORECASE,
)


def _clean_spaces(s: str) -> str:
    s = s.replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def _lines(text: str) -> List[str]:
    text = _clean_spaces(text)
    lines = [ln.strip() for ln in text.split("\n")]
    return [ln for ln in lines if ln.strip()]


def parse_uy_amount(raw: Optional[str]) -> Optional[float]:
    """
    Convierte montos típicos UY:
    - "8.516,99" -> 8516.99
    - "4 018,01" -> 4018.01
    - "0,00" -> 0.0
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None

    s = s.replace(" ", "")
    s = re.sub(r"[^0-9,\.\-]", "", s)

    if not s or s in {"-", ".", ",", "-.", "-,"}:
        return None

    # Si hay coma, la interpretamos como decimal (última coma)
    if "," in s:
        parts = s.split(",")
        dec = parts[-1]
        intpart = "".join(parts[:-1]).replace(".", "")
        if not intpart or intpart in {"-", "+"}:
            return None
        try:
            return float(intpart + "." + dec)
        except ValueError:
            return None

    # Sin coma: puede venir con miles por puntos
    if s.count(".") > 1:
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def fmt_uy_amount(num: Optional[float]) -> Optional[str]:
    if num is None:
        return None
    # Siempre 2 decimales, coma decimal
    return f"{num:.2f}".replace(".", ",")


def looks_like_person_name(line: str) -> bool:
    """
    Heurística: muchos PDFs meten un nombre de persona suelto antes del RUT del receptor.
    Si tiene 3+ palabras y no tiene pinta de empresa, lo consideramos "persona/titular",
    NO razón social del emisor.
    """
    words = [w for w in line.split() if w.strip()]
    if len(words) >= 3 and not BUSINESS_HINT_RE.search(line):
        return True
    return False


def is_probably_header(line: str) -> bool:
    return bool(re.search(r"\b(RUT|RUC)\b|\bTIPO DOCUMENTO\b|\bSERIE\b|\bNUMERO\b|\bFORMA DE PAGO\b", line, re.I))


def block_after_label(
    lines: List[str],
    label_regex: str,
    end_regexes: Optional[List[str]] = None,
    max_lines: int = 20,
) -> List[str]:
    """
    Devuelve un bloque de líneas después de encontrar un label que matchee label_regex.
    Se detiene si encuentra cualquiera de end_regexes.
    """
    idx = None
    for i, ln in enumerate(lines):
        if re.search(label_regex, ln, re.I):
            idx = i
            break
    if idx is None:
        return []

    out: List[str] = []
    for j in range(idx + 1, min(len(lines), idx + 1 + max_lines)):
        ln = lines[j]
        if end_regexes and any(re.search(er, ln, re.I) for er in end_regexes):
            break
        out.append(ln)
    return out


# -----------------------------
# Modelo de salida
# -----------------------------

@dataclass
class InvoiceRecord:
    fecha: Optional[str] = None
    serie: Optional[str] = None
    folio: Optional[str] = None
    serie_y_folio: Optional[str] = None

    razon_social: Optional[str] = None  # Emisor (si se puede inferir)
    razon_social_receptor: Optional[str] = None

    rut_emisor: Optional[str] = None
    rut_receptor: Optional[str] = None

    tipo_documento: Optional[str] = None
    forma_pago: Optional[str] = None
    vencimiento: Optional[str] = None
    moneda: Optional[str] = None

    importe_total_con_iva: Optional[str] = None
    importe_total_con_iva_num: Optional[float] = None

    importe_sin_iva: Optional[str] = None
    importe_sin_iva_num: Optional[float] = None
    importe_sin_iva_fuente: Optional[str] = None

    es_nota_de_credito: bool = False

    _archivo: Optional[str] = None
    _fuente: Optional[str] = None
    _debug: Optional[str] = None


# -----------------------------
# PDF parsing
# -----------------------------

def extract_pdf_text(pdf_path: Path) -> str:
    try:
        import pdfplumber  # type: ignore
    except Exception as e:
        raise RuntimeError("Falta dependencia: pdfplumber. Instalá: pip install pdfplumber") from e

    parts: List[str] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            parts.append(txt)
    return _clean_spaces("\n\n".join(parts))


def parse_emisor_rut_y_tipo(lines: List[str]) -> Tuple[Optional[str], Optional[str]]:
    joined = "\n".join(lines)
    # Caso típico: "RUT EMISOR TIPO DOCUMENTO" y en la siguiente línea: "210... e-Factura"
    m = re.search(r"RUT\s+EMISOR\s+TIPO\s+DOCUMENTO\s+([0-9]{12})\s+(.+)", joined, re.I)
    if m:
        rut = m.group(1).strip()
        td = m.group(2).strip()
        td = re.sub(r"\s{2,}", " ", td)
        return rut, td

    # Fallback: buscar una línea con rut cerca de esa cabecera
    for i, ln in enumerate(lines):
        if re.search(r"RUT\s+EMISOR\s+TIPO\s+DOCUMENTO", ln, re.I):
            for j in range(i + 1, min(i + 6, len(lines))):
                m2 = re.search(r"\b([0-9]{12})\b\s+(.+)", lines[j])
                if m2:
                    return m2.group(1).strip(), m2.group(2).strip()
    return None, None


def parse_serie_folio_forma_pago_venc(lines: List[str]) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[str]]:
    """
    Busca el bloque:
    SERIE NUMERO FORMA DE PAGO VENCIMIENTO
    A 133716 Credito 01/12/2025
    """
    block = block_after_label(
        lines,
        r"SERIE\s+NUMERO\s+FORMA\s+DE\s+PAGO",
        end_regexes=[r"RUT\s+RECEPTOR", r"RAZON\s+SOCIAL", r"DIRECCION", r"PAIS\s+FECHA"],
        max_lines=6,
    )
    line = block[0] if block else ""

    if not line:
        return None, None, None, None, None

    # Captura serie (1 letra), folio (número), y el resto
    m = re.search(r"\b([A-Z])\b\s*0*([0-9]{1,12})\b(.*)$", line)
    if not m:
        return None, None, None, None, line.strip() or None

    serie = m.group(1).strip()
    folio = str(int(m.group(2)))
    rest = (m.group(3) or "").strip()

    vencimiento = None
    dm = UY_DATE_RE.search(rest)
    if dm:
        vencimiento = dm.group(1)
        rest = rest.replace(dm.group(1), "").strip()

    forma_pago = rest or None
    serie_y_folio = f"{serie}-{folio}" if serie and folio else None
    return serie, folio, serie_y_folio, forma_pago, vencimiento


def parse_receptor(lines: List[str]) -> Tuple[Optional[str], Optional[str], Optional[str], Dict[str, Any]]:
    """
    Intenta obtener:
    - rut_receptor
    - razon_social_receptor
    - razon_social (emisor) si aparece suelta (como "Cafe Bahia") en algunos PDFs
    """
    dbg: Dict[str, Any] = {}

    block = block_after_label(
        lines,
        r"RUT\s+RECEPTOR\s+RAZON\s+SOCIAL",
        end_regexes=[r"DIRECCION", r"PAIS\s+FECHA", r"CONCEPTO", r"SUBTOTAL", r"TOTAL\s+A\s+PAGAR"],
        max_lines=12,
    )
    dbg["rr_block"] = block[:]

    rut_receptor, razon_receptor = None, None
    pre_lines: List[str] = []

    for ln in block:
        m = re.search(r"\b([0-9]{12})\b\s*(.*)$", ln)
        if m:
            rut_receptor = m.group(1).strip()
            tail = (m.group(2) or "").strip()
            if tail:
                razon_receptor = tail
            break
        pre_lines.append(ln)

    # Si el receptor no viene con nombre en la línea del rut, a veces lo pusieron justo antes.
    if razon_receptor is None and pre_lines:
        # El último "pre" suele ser el nombre si está.
        cand = pre_lines[-1].strip()
        if cand and not is_probably_header(cand):
            razon_receptor = cand

    razon_emisor = None
    if pre_lines:
        # Caso típico (como tu "Cafe Bahia"):
        # Si hay un renglón suelto ANTES del rut del receptor y ES CORTO (<=2 palabras),
        # lo tratamos como posible emisor (no perfecto, pero evita "ERICA AIDEE ...").
        cand = pre_lines[0].strip()
        if cand and not looks_like_person_name(cand) and len(cand.split()) <= 2:
            razon_emisor = cand

    dbg["rr_pre_lines"] = pre_lines[:]
    return rut_receptor, razon_receptor, razon_emisor, dbg


def parse_pais_fecha_moneda(lines: List[str]) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    block = block_after_label(
        lines,
        r"PAIS\s+FECHA\s+DE\s+DOCUMENTO\s+MONEDA",
        end_regexes=[r"CONCEPTO", r"DESC", r"SUBTOTAL", r"TOTAL\s+A\s+PAGAR"],
        max_lines=4,
    )
    line = block[0] if block else ""
    if not line:
        return None, None, None

    # Ej: "UY 26/11/2025 Peso Uruguayo"
    m = re.search(r"\b([A-Z]{2})\b\s*(\d{1,2}/\d{1,2}/\d{4})\s*(.+)$", line)
    if m:
        pais = m.group(1).strip()
        fecha = m.group(2).strip()
        moneda = m.group(3).strip() or None
        return pais, fecha, moneda

    # Fallback: buscar fecha en esa línea
    dm = UY_DATE_RE.search(line)
    if dm:
        return None, dm.group(1), None

    return None, None, None


def last_amount(joined: str, label_regex: str) -> Optional[str]:
    # Captura montos tipo "1.050,81" o "4 018,01"
    pattern = re.compile(label_regex + r"\s*([0-9][0-9\.\s]*,[0-9]{2})", re.I)
    matches = pattern.findall(joined)
    return matches[-1] if matches else None


def parse_totales(lines: List[str]) -> Tuple[Optional[str], Optional[float], Optional[float], Optional[str], Dict[str, Any]]:
    joined = "\n".join(lines)
    dbg: Dict[str, Any] = {}

    total_str = None
    # "Total a pagar 8.516,99"
    total_matches = re.findall(r"Total\s+a\s+pagar\s+([0-9][0-9\.\s]*,[0-9]{2})", joined, re.I)
    if total_matches:
        total_str = total_matches[-1]

    # Fallback: algunos ponen "TOTAL 4 018,01"
    if not total_str:
        total_matches2 = re.findall(r"\bTOTAL\b\s*([0-9][0-9\.\s]*,[0-9]{2})", joined, re.I)
        if total_matches2:
            # Elegimos el más grande para evitar agarrar cualquier "TOTAL 0,00"
            vals = [(m, parse_uy_amount(m) or 0.0) for m in total_matches2]
            vals.sort(key=lambda x: x[1])
            total_str = vals[-1][0] if vals else None

    total_num = parse_uy_amount(total_str) if total_str else None

    sub10 = last_amount(joined, r"Subtotal\s+gravado\s*\(10%\)")
    iva10 = last_amount(joined, r"Total\s+iva\s*\(10%\)")
    sub22 = last_amount(joined, r"Subtotal\s+gravado\s*\(22%\)")
    iva22 = last_amount(joined, r"Total\s+iva\s*\(22%\)")
    sub_ng = last_amount(joined, r"Subtotal\s+no\s+gravado")

    dbg.update(
        {
            "sub10": sub10,
            "iva10": iva10,
            "sub22": sub22,
            "iva22": iva22,
            "sub_no_gravado": sub_ng,
        }
    )

    # Mejor fuente para neto: subtotales (si existen)
    net_parts = [parse_uy_amount(x) for x in (sub10, sub22, sub_ng)]
    net_num = None
    net_src = None

    if any(v is not None for v in net_parts):
        net_num = sum(v for v in net_parts if v is not None)
        net_src = "subtotales"
    else:
        # Si tenemos total + IVA(s), neto = total - iva_total
        iva_parts = [parse_uy_amount(x) for x in (iva10, iva22)]
        iva_total = None
        if any(v is not None for v in iva_parts):
            iva_total = sum(v for v in iva_parts if v is not None)
        if total_num is not None and iva_total is not None:
            net_num = total_num - iva_total
            net_src = "total_menos_iva"

    return total_str, total_num, net_num, net_src, dbg


def guess_emisor_name(lines: List[str]) -> Optional[str]:
    """
    Emisor: mejor esfuerzo.
    1) Si el PDF trae un nombre suelto antes de "RUT EMISOR", lo tomamos.
    2) Si no hay, queda None y listo (mejor None que inventar basura).
    """
    for ln in lines[:8]:
        if re.search(r"RUT\s+EMISOR", ln, re.I):
            break
        if ln and not is_probably_header(ln) and not re.search(r"Página|OBSERVACIONES|Codigo de Seguridad", ln, re.I):
            return ln.strip()
    return None


def parse_pdf_invoice(pdf_path: Path, debug: bool = False) -> InvoiceRecord:
    text = extract_pdf_text(pdf_path)
    lines = _lines(text)
    joined = "\n".join(lines)

    rut_emisor, tipo_doc = parse_emisor_rut_y_tipo(lines)
    serie, folio, serie_y_folio, forma_pago, vencimiento = parse_serie_folio_forma_pago_venc(lines)

    rut_receptor, razon_receptor, razon_emisor_hint, rr_dbg = parse_receptor(lines)
    _pais, fecha, moneda = parse_pais_fecha_moneda(lines)

    total_str, total_num, net_num, net_src, tot_dbg = parse_totales(lines)

    razon_emisor = guess_emisor_name(lines) or razon_emisor_hint

    # Nota de crédito: detectar por tipo documento o texto (no por "Credito" de forma de pago)
    es_nc = False
    if tipo_doc and re.search(r"nota\s+de\s+cr[eé]dito|n[./]c|e-?nota", tipo_doc, re.I):
        es_nc = True
    elif re.search(r"nota\s+de\s+cr[eé]dito", joined, re.I):
        es_nc = True

    rec = InvoiceRecord(
        fecha=fecha,
        serie=serie,
        folio=folio,
        serie_y_folio=serie_y_folio,
        razon_social=razon_emisor,
        razon_social_receptor=razon_receptor,
        rut_emisor=rut_emisor,
        rut_receptor=rut_receptor,
        tipo_documento=tipo_doc,
        forma_pago=forma_pago,
        vencimiento=vencimiento,
        moneda=moneda,
        importe_total_con_iva=total_str,
        importe_total_con_iva_num=total_num,
        importe_sin_iva=fmt_uy_amount(net_num) if net_num is not None else None,
        importe_sin_iva_num=net_num,
        importe_sin_iva_fuente=net_src,
        es_nota_de_credito=es_nc,
        _archivo=str(pdf_path),
        _fuente="pdf_text",
        _debug=None,
    )

    if debug:
        dbg = {
            "rut_emisor": rut_emisor,
            "tipo_documento": tipo_doc,
            "serie_folio_line": {"serie": serie, "folio": folio, "forma_pago": forma_pago, "vencimiento": vencimiento},
            "receptor": rr_dbg,
            "pais_fecha_moneda": {"pais": _pais, "fecha": fecha, "moneda": moneda},
            "totales": tot_dbg,
        }
        rec._debug = json.dumps(dbg, ensure_ascii=False)

    return rec


# -----------------------------
# OCR (opcional, lazy import)
# -----------------------------

def apply_resource_limits(cpu_threads: int = 1, low_priority: bool = False) -> None:
    """
    Limita hilos tanto como podamos (sin prometer magia).
    Importante: se ejecuta antes de cargar cosas pesadas (torch/opencv).
    """
    cpu_threads = max(1, int(cpu_threads))

    # OpenMP / MKL (numpy/torch)
    os.environ.setdefault("OMP_NUM_THREADS", str(cpu_threads))
    os.environ.setdefault("MKL_NUM_THREADS", str(cpu_threads))
    os.environ.setdefault("NUMEXPR_NUM_THREADS", str(cpu_threads))
    os.environ.setdefault("OPENBLAS_NUM_THREADS", str(cpu_threads))
    os.environ.setdefault("VECLIB_MAXIMUM_THREADS", str(cpu_threads))

    if low_priority and os.name == "nt":
        try:
            import ctypes  # noqa
            import ctypes.wintypes  # noqa

            PROCESS_SET_INFORMATION = 0x0200
            IDLE_PRIORITY_CLASS = 0x00000040

            kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
            GetCurrentProcess = kernel32.GetCurrentProcess
            SetPriorityClass = kernel32.SetPriorityClass

            hproc = GetCurrentProcess()
            SetPriorityClass(hproc, IDLE_PRIORITY_CLASS)
        except Exception:
            # Si no se puede, seguimos igual.
            pass


def parse_image_invoice_with_easyocr(
    image_path: Path,
    debug: bool = False,
    max_dim: int = 1800,
    max_pixels: int = 2_500_000,
) -> InvoiceRecord:
    """
    OCR medio "best-effort". No lo uses si podés conseguir PDFs con texto.
    """
    try:
        import cv2  # type: ignore
        import easyocr  # type: ignore
    except Exception as e:
        raise RuntimeError("Para OCR necesitás: pip install opencv-python easyocr") from e

    img = cv2.imread(str(image_path))
    if img is None:
        raise RuntimeError("No se pudo leer la imagen (cv2.imread devolvió None).")

    h, w = img.shape[:2]
    # Limit pixels by scaling
    scale = 1.0
    if w > max_dim or h > max_dim:
        scale = min(max_dim / float(w), max_dim / float(h))
    if (w * h) > max_pixels:
        scale2 = (max_pixels / float(w * h)) ** 0.5
        scale = min(scale, scale2)
    if scale < 1.0:
        img = cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # Umbral suave para ayudar
    gray = cv2.bilateralFilter(gray, 5, 60, 60)
    _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    reader = easyocr.Reader(["es"], gpu=False)
    result = reader.readtext(thr, detail=0)
    text = _clean_spaces("\n".join(result))
    lines = _lines(text)

    # Intentamos reutilizar lógica de PDF sobre texto OCR
    rut_emisor, tipo_doc = parse_emisor_rut_y_tipo(lines)
    serie, folio, serie_y_folio, forma_pago, vencimiento = parse_serie_folio_forma_pago_venc(lines)
    rut_receptor, razon_receptor, razon_emisor_hint, rr_dbg = parse_receptor(lines)
    _pais, fecha, moneda = parse_pais_fecha_moneda(lines)
    total_str, total_num, net_num, net_src, tot_dbg = parse_totales(lines)

    razon_emisor = guess_emisor_name(lines) or razon_emisor_hint
    es_nc = False
    if tipo_doc and re.search(r"nota\s+de\s+cr[eé]dito|n[./]c|e-?nota", tipo_doc, re.I):
        es_nc = True
    elif re.search(r"nota\s+de\s+cr[eé]dito", "\n".join(lines), re.I):
        es_nc = True

    rec = InvoiceRecord(
        fecha=fecha,
        serie=serie,
        folio=folio,
        serie_y_folio=serie_y_folio,
        razon_social=razon_emisor,
        razon_social_receptor=razon_receptor,
        rut_emisor=rut_emisor,
        rut_receptor=rut_receptor,
        tipo_documento=tipo_doc,
        forma_pago=forma_pago,
        vencimiento=vencimiento,
        moneda=moneda,
        importe_total_con_iva=total_str,
        importe_total_con_iva_num=total_num,
        importe_sin_iva=fmt_uy_amount(net_num) if net_num is not None else None,
        importe_sin_iva_num=net_num,
        importe_sin_iva_fuente=net_src,
        es_nota_de_credito=es_nc,
        _archivo=str(image_path),
        _fuente="image_ocr_easyocr",
        _debug=None,
    )

    if debug:
        dbg = {
            "ocr_text_preview": text[:1500],
            "receptor": rr_dbg,
            "totales": tot_dbg,
        }
        rec._debug = json.dumps(dbg, ensure_ascii=False)
    return rec


# -----------------------------
# IO: outputs
# -----------------------------

OUTPUT_COLUMNS = [
    "fecha",
    "serie",
    "folio",
    "serie_y_folio",
    "razon_social",
    "razon_social_receptor",
    "rut_emisor",
    "rut_receptor",
    "tipo_documento",
    "forma_pago",
    "vencimiento",
    "moneda",
    "importe_total_con_iva",
    "importe_total_con_iva_num",
    "importe_sin_iva",
    "importe_sin_iva_num",
    "importe_sin_iva_fuente",
    "es_nota_de_credito",
    "_archivo",
    "_fuente",
    "_debug",
]


def write_csv(path: Path, records: List[InvoiceRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        w.writeheader()
        for r in records:
            d = asdict(r)
            w.writerow({k: d.get(k) for k in OUTPUT_COLUMNS})


def write_xlsx(path: Path, records: List[InvoiceRecord]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception as e:
        raise RuntimeError("Para XLSX necesitás: pip install openpyxl") from e

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "facturas"

    ws.append(OUTPUT_COLUMNS)
    for r in records:
        d = asdict(r)
        ws.append([d.get(k) for k in OUTPUT_COLUMNS])

    # Autowidth simple (sin volverse loco)
    for i, col in enumerate(OUTPUT_COLUMNS, start=1):
        max_len = len(col)
        for row in ws.iter_rows(min_col=i, max_col=i, min_row=2, values_only=True):
            v = row[0]
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

    wb.save(str(path))


# -----------------------------
# Gold report
# -----------------------------

def load_gold(path: Path) -> Dict[str, Dict[str, Any]]:
    """
    Soporta gold.json como:
    [
      {"archivo": "factura.pdf", "fecha": "...", "rut_emisor": "...", ...},
      ...
    ]
    Indexado por basename.
    """
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("gold.json debe ser una LISTA de objetos.")
    out: Dict[str, Dict[str, Any]] = {}
    for item in data:
        if not isinstance(item, dict):
            continue
        name = item.get("archivo") or item.get("_archivo") or item.get("file")
        if not name:
            continue
        out[Path(name).name] = item
    return out


def report_against_gold(records: List[InvoiceRecord], gold_map: Dict[str, Dict[str, Any]]) -> None:
    # Campos comparables (los importantes)
    fields = ["fecha", "rut_emisor", "serie", "folio", "importe_total_con_iva_num"]

    covered = 0
    correct = {f: 0 for f in fields}

    mismatches: List[str] = []

    for r in records:
        base = Path(r._archivo or "").name
        g = gold_map.get(base)
        if not g:
            continue
        covered += 1
        rd = asdict(r)

        for f in fields:
            pred = rd.get(f)
            gold = g.get(f)
            ok = False

            if f == "importe_total_con_iva_num":
                # Comparación float con tolerancia chica
                try:
                    ok = (pred is not None and gold is not None and abs(float(pred) - float(gold)) < 0.01)
                except Exception:
                    ok = False
            else:
                ok = (pred == gold)

            if ok:
                correct[f] += 1
            else:
                mismatches.append(f"* {base}\n  - {f}: pred={pred!r} gold={gold!r}")

    print("\n=== REPORT ===")
    print(f"Docs con gold: {covered}")
    if covered == 0:
        return
    for f in fields:
        pct = 100.0 * correct[f] / covered
        print(f"{f}: {correct[f]}/{covered} ({pct:.1f}%)")

    if mismatches:
        print("\n--- MISMATCHES ---")
        for m in mismatches:
            print(m)


# -----------------------------
# Main
# -----------------------------

def iter_files(root: Path, recursive: bool = True) -> Iterable[Path]:
    if root.is_file():
        yield root
        return
    if recursive:
        for p in root.rglob("*"):
            if p.is_file():
                yield p
    else:
        for p in root.glob("*"):
            if p.is_file():
                yield p


def main() -> int:
    ap = argparse.ArgumentParser(description="Parseador de facturas (PDF-text primero, OCR opcional).")
    ap.add_argument("input", help="Carpeta o archivo a procesar.")
    ap.add_argument("--json", action="store_true", help="Imprime JSON a stdout.")
    ap.add_argument("--csv", help="Ruta de salida CSV.")
    ap.add_argument("--xlsx", help="Ruta de salida XLSX.")
    ap.add_argument("--debug", action="store_true", help="Incluye _debug con info de extracción.")
    ap.add_argument("--recursive", action="store_true", help="Buscar recursivamente en subcarpetas.")
    ap.add_argument("--no-ocr", action="store_true", help="Desactiva OCR (recomendado si usás PDFs con texto).")

    # Controles de recursos (principalmente para OCR)
    ap.add_argument("--cpu-threads", type=int, default=1, help="Límite de hilos (mejor 1 si tu PC muere).")
    ap.add_argument("--low-priority", action="store_true", help="Baja prioridad del proceso (Windows).")
    ap.add_argument("--max-dim", type=int, default=1800, help="Máx dimensión para OCR (resize).")
    ap.add_argument("--max-pixels", type=int, default=2_500_000, help="Máx pixeles para OCR (resize).")

    # Dedupe
    ap.add_argument("--dedup", action="store_true", help="Elimina duplicados por (rut_emisor, serie, folio, total).")

    # Report
    ap.add_argument("--report", action="store_true", help="Genera reporte comparando con gold.json.")
    ap.add_argument("--gold", help="Ruta a gold.json (lista de objetos con 'archivo' y campos).")

    args = ap.parse_args()

    apply_resource_limits(cpu_threads=args.cpu_threads, low_priority=args.low_priority)

    root = Path(args.input).expanduser()
    if not root.exists():
        print(f"[ERROR] No existe: {root}", file=sys.stderr)
        return 2

    records: List[InvoiceRecord] = []
    seen: set = set()

    for f in iter_files(root, recursive=args.recursive):
        ext = f.suffix.lower()
        try:
            if ext == ".pdf":
                rec = parse_pdf_invoice(f, debug=args.debug)
            elif ext in {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".webp"}:
                if args.no_ocr:
                    # Saltamos imágenes si no queremos OCR
                    continue
                rec = parse_image_invoice_with_easyocr(
                    f,
                    debug=args.debug,
                    max_dim=args.max_dim,
                    max_pixels=args.max_pixels,
                )
            else:
                continue

            # Fix serie_y_folio si faltaba
            if rec.serie and rec.folio and not rec.serie_y_folio:
                rec.serie_y_folio = f"{rec.serie}-{rec.folio}"

            # Dedupe
            if args.dedup:
                key = (
                    rec.rut_emisor or "",
                    rec.serie or "",
                    rec.folio or "",
                    f"{rec.importe_total_con_iva_num or ''}",
                )
                if key in seen:
                    continue
                seen.add(key)

            records.append(rec)

        except Exception as e:
            # Un error en un archivo no debería matar el lote
            err = InvoiceRecord(
                razon_social=None,
                _archivo=str(f),
                _fuente="error",
                _debug=str(e) if args.debug else None,
            )
            records.append(err)

    # Outputs
    if args.csv:
        write_csv(Path(args.csv), records)

    if args.xlsx:
        write_xlsx(Path(args.xlsx), records)

    if args.json or (not args.csv and not args.xlsx and not args.report):
        print(json.dumps([asdict(r) for r in records], ensure_ascii=False, indent=2))

    # Report
    if args.report:
        if not args.gold:
            print("[ERROR] --report requiere --gold ruta/al/gold.json", file=sys.stderr)
            return 2
        gold_map = load_gold(Path(args.gold))
        report_against_gold(records, gold_map)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
